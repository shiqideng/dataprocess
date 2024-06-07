# 标准库
import sys
import os
import re
import glob
import csv
import openpyxl
from configparser import ConfigParser

# 第三方库
from PySide6.QtWidgets import QWidget, QFileDialog, QHeaderView
from PySide6.QtCore import QStandardPaths, Slot, Qt
import openpyxl.writer

# 本地包
from UI.Ui_MainWindow import Ui_Form
from DataOperation import Module


class MainWindow(QWidget, Ui_Form):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.run()

    def run(self):
        # 初始化配置
        self.getSystem()
        self.DataProcess5400ProgressBar.setValue(0)
        # 隐藏peaktable标签页
        self.QTabWidget5400.removeTab(2)
        # 隐藏labchip标签页
        self.QTabWidget5400.removeTab(2)
        self.ReName5400CheckBox.setChecked(False)
        self.ReName5400CheckBox.setDisabled(True)
        self.Waring5400Label.setVisible(False)
        self.saveName = ""

        # 使用lambda函数来传递参数
        self.Import5400FilePathToolButton.clicked.connect(lambda: self.selectFile("Import5400FilePathLineEdit"))
        self.Export5400FilePathToolButton.clicked.connect(lambda: self.selectFile("Export5400FilePathLineEdit"))
        self.Preview5400PushButton.clicked.connect(self.preview5400)
        self.Start5400PushButton.clicked.connect(self.start5400)
        self.Export5400PushButton.clicked.connect(self.exportToCSV)
        self.Clear5400PushButton.clicked.connect(self.clear5400)
        self.SampleType5400ComboBox.currentTextChanged.connect(self.handleIndexChanged)

    @Slot(str)
    def selectFile(self, lineEditName):
        defaultPath = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
        filePath = QFileDialog.getExistingDirectory(self, "选取文件夹", defaultPath)
        if filePath:
            getattr(self, lineEditName).setText(filePath)
        else:
            self.logTextBrowser.append(Module.logFormat("INFO" ,"用户取消了文件夹选择操作或出现了错误。"))
    
    def exportToCSV(self):
        CELLTOCLEAR=["备注", "空列", "原始质量浓度", "原始摩尔浓度"]
        MERGERANGE="G1:H1"
        HEADTITLE = "片段大小"
        fileName = self.Export5400FilePathLineEdit.text() + "\\" + self.saveName + ".csv"
        # self.Export5400FilePathLineEdit.setText(fileName)
        model = self.ResultTableAgilent5400TableView.model()
        columnCount = model.columnCount()
        rowCount = model.rowCount()
        if not os.path.exists(self.Export5400FilePathLineEdit.text()):
            os.makedirs(self.Export5400FilePathLineEdit.text())
        with open(fileName, 'w', newline='') as file:
            writer = csv.writer(file)
            # 写入表头
            headerData = [model.headerData(column, Qt.Horizontal, Qt.DisplayRole) for column in range(columnCount)]
            writer.writerow(headerData)

            for row in range(rowCount):
                rowData = []
                for column in range(columnCount):
                    index = model.index(row, column)
                    value = model.data(index, Qt.DisplayRole)
                    rowData.append(value)
                writer.writerow(rowData)
        try:  
            # 加载并处理Excel文件

            with open(fileName, 'r') as file:
                reader = csv.reader(file)
                data = list(reader)
            
            # 创建Excel工作簿对象
            wb = openpyxl.Workbook()
            sheet = wb.active
            
            # 将CSV数据逐行写入Excel工作表
            for row in data:
                sheet.append(row)
            
            # 高效地清空指定单元格内容
            for cell_value in CELLTOCLEAR:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value == cell_value:
                            cell.value = None

            # 合并单元格
            sheet.merge_cells(MERGERANGE)
            sheet[MERGERANGE.split(":")[0]] = HEADTITLE

            # 保存修改
            wb.save(fileName.split(".")[0]+ ".xlsx")
            wb = openpyxl.load_workbook(fileName.split(".")[0]+ ".xlsx")
            sheet = wb.active
            for col in 'GHI':
                for row in sheet[col]:
                    if row.value == 0:
                        row.value = None  # 将值为0的单元格设置为None（清空内容）
            wb.save(fileName.split(".")[0]+ ".xlsx")
            os.remove(fileName)
        except Exception as e:
            print(f"Error occurred: {e}")
            # 根据实际情况，可以选择重新抛出异常或者进行其他错误处理

        self.logTextBrowser.append(Module.logFormat("INFO", f'数据已成功导出到{fileName}!'))
        self.messagebox = Module.MessageBox(Icon="Information", text=f'数据已成功导出到{fileName}!')
        self.messagebox.show()

    def handleIndexChanged(self):
        if self.SampleType5400ComboBox.currentText() == "核酸":
            self.QTabWidget5400.insertTab(2, self.PeakTable5400tab, "Peak Table")
        elif self.SampleType5400ComboBox.currentText() == "文库":
            self.QTabWidget5400.removeTab(2)

    def start5400(self):
        if self.Import5400FilePathLineEdit.text():
            try:
                self.preview5400()
                self.upDatePrograssBar(15)
                filePath = findFilePath({"path": self.Import5400FilePathLineEdit.text(), "SampleType":self.SampleType5400ComboBox.currentText()})
                filePath = filePath.getFilePath()
                if filePath["reg"] == 1:
                    self.saveName = filePath["msg"]["saveName"]
                    if self.SampleType5400ComboBox.currentText() == "核酸":
                        pass
                    elif self.SampleType5400ComboBox.currentText() == "文库":
                        global qualityTablepath
                        qualityTablepath = filePath["msg"]["qualityTable"]
                        global smearTablepath
                        smearTablepath = filePath["msg"]["smearTable"]
                        filePath = {"SampleType":"文库", "qualityTablepath": qualityTablepath, "smearTablepath":smearTablepath}
                        self.worker = Module.caculateResult(filePath)
                        if self.worker["reg"] == 1:
                            # ResultModule = self.worker.prograssChange.connect(self.upDatePrograssBar)
                            Result = Module.createModel({"type": 1, "path":self.worker["msg"]})
                            # self.ResultTableLabChip5400TableView.setModel(LabChipResult)
                            self.ResultTableAgilent5400TableView.setModel(Result)
                        # 设置表格行和列自适应
                        self.ResultTableLabChip5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableLabChip5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableLabChip5400TableView.resizeColumnsToContents()

                        self.ResultTableAgilent5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableAgilent5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableAgilent5400TableView.resizeColumnsToContents()
                else:
                    self.logTextBrowser.append(Module.logFormat("ERROR", str(filePath["msg"])))
            except IOError:
                self.logTextBrowser.append(Module.logFormat("ERROR", "部分文件存在异常"+ str(IOError)))
            except Exception as e:
                self.logTextBrowser.append(Module.logFormat("ERROR", "发生意外错误"+ str(e)))
        else:
            self.logTextBrowser.append(Module.logFormat("ERROR", "请选择导入文件路径。"))
            
    def clear5400(self):
        self.messagebox = Module.MessageBox(Icon="Question", text="确定清空所有数据吗？")
        self.messagebox.resultReady.connect(self.handleResultReady)
        self.messagebox.show()

    def clearData(self):
        # 定义需要清除数据的表格视图列表
        tablesToClear = [
            self.PeakTable5400TableView,
            self.SmearTable5400TableView,
            self.QualityTable5400TableView,
            self.ResultTableAgilent5400TableView
        ]
        
        # 清除文本框中的数据
        self.Export5400FilePathLineEdit.clear()
        self.Import5400FilePathLineEdit.clear()

        # 遍历表格视图列表，尝试清除每个表格的数据
        for tableView in tablesToClear:
            # 为了提高代码健壮性，添加了对table_view对象存在性的检查
            if tableView is not None:
                # 再次检查模型是否存在，如果存在，则清除模型
                model = tableView.model()
                if model is not None:
                    tableView.setModel(None)
            else:
                # 可以选择在这里记录日志或者进行一些错误处理
                self.logTextBrowser.append(format("ERROR"), f"Warning: Table view {tableView} is None.")

                
    def preview5400(self):
        if self.Import5400FilePathLineEdit.text():
            try:
                filePath = findFilePath({"path": self.Import5400FilePathLineEdit.text(), "SampleType":self.SampleType5400ComboBox.currentText()})
                filePath = filePath.getFilePath()
                if filePath["reg"] == 1:
                    self.saveName = filePath["msg"]["saveName"]
                    if self.SampleType5400ComboBox.currentText() == "核酸":
                        self.logTextBrowser.append(Module.logFormat("INFO", "核酸分析暂未支持，请等待后续升级！"))
                        self.messageBox = Module.MessageBox(Icon="Warning", text="核酸结果分析暂未支持，请等待后续升级！")
                        self.messageBox.show()

                        # self.logTextBrowser.append(Module.logFormat("INFO", "qualityTablePath:" + str(filePath["msg"]["qualityTable"])))
                        # self.logTextBrowser.append(Module.logFormat("INFO", "smearTablePath:" + str(filePath["msg"]["smearTable"])))
                        # self.logTextBrowser.append(Module.logFormat("INFO", "peakTablePath:" + str(filePath["msg"]["peakTable"])))
                        
                        # global qualityTablepath
                        # qualityTablePath = filePath["msg"]["qualityTable"]
                        # quality = Module.createModel(qualityTablePath)
                        # self.QualityTable5400TableView.setModel(quality)
                        # # 设置表格行和列自适应
                        # self.QualityTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.QualityTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                        # global smearTablepath
                        # smearTablePath = filePath["msg"]["smearTable"]
                        # smear = Module.createModel(smearTablePath)
                        # self.SmearTable5400TableView.setModel(smear)
                        # # 设置表格行和列自适应
                        # self.SmearTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.SmearTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                        # global peakTablePath
                        # peakTablePath = filePath["msg"]["peakTable"]
                        # peak =Module.createModel(peakTablePath)
                        # self.PeakTable5400TableView.setModel(peak)
                        # # 设置表格行和列自适应
                        # self.PeakTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.PeakTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                    elif self.SampleType5400ComboBox.currentText() == "文库":
                        self.logTextBrowser.append(Module.logFormat("INFO", "qualityTablePath:" + str(filePath["msg"]["qualityTable"])))
                        self.logTextBrowser.append(Module.logFormat("INFO", "smearTablePath:" + str(filePath["msg"]["smearTable"])))

                        global qualityTablepath
                        qualityTablepath = filePath["msg"]["qualityTable"]
                        quality = Module.createModel({"type": 2, "path":qualityTablepath})
                        self.QualityTable5400TableView.setModel(quality)
                        # 设置表格行和列自适应
                        self.QualityTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.QualityTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.QualityTable5400TableView.resizeColumnsToContents()
                        
                        global smearTablepath
                        smearTablepath = filePath["msg"]["smearTable"]
                        smear = Module.createModel({"type": 2, "path":smearTablepath})
                        self.SmearTable5400TableView.setModel(smear)
                        # 设置表格行和列自适应
                        self.SmearTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.SmearTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                        
                else:
                    self.logTextBrowser.append(Module.logFormat("ERROR", str(filePath["msg"])))
            except IOError:
                self.logTextBrowser.append(Module.logFormat("ERROR", "部分文件存在异常"+ str(IOError)))
            except Exception as e:
                self.logTextBrowser.append(Module.logFormat("ERROR", "发生意外错误"+ str(e)))
        else:
            self.logTextBrowser.append(Module.logFormat("ERROR", "请选择导入文件路径。"))


    
    def getSystem(self):
        try:
            if sys.platform == 'linux' or sys.platform == 'darwin':
                importPath = Module.getConfig('config.ini','Import Paths', 'linux_path')
                exportPath = Module.getConfig('config.ini','Export Paths', 'linux_path')
            elif sys.platform == 'win32':
                importPath = Module.getConfig('config.ini','Import Paths', 'win32_path')
                exportPath = Module.getConfig('config.ini','Export Paths', 'win32_path')
            # 使用 os.path.join 来构造路径，即使这里的路径已经是完整的
            importPath = os.path.join(importPath["msg"])  # 为了演示，实际上这里并不需要join
            exportPath = os.path.join(exportPath["msg"])
            self.Import5400FilePathLineEdit.setText(importPath)
            self.Export5400FilePathLineEdit.setText(exportPath)
        except Exception as e:
            # 在这里处理可能的异常，比如读取配置文件失败、路径不存在等
            self.logTextBrowser.append(Module.logFormat("ERROR", str(e)))
    
    def upDatePrograssBar(self, value):
        self.DataProcess5400ProgressBar.setValue(value)
    
    @Slot(dict)
    def handleResultReady(self, message):
        if message["reg"] == 1:
            self.logTextBrowser.append(Module.logFormat("INFO", "用户确认清除内容" ))
            self.clearData()
            # self.logOperationSuccess()
        else:
            # self.logOperationCancelled()
            self.logTextBrowser.append(Module.logFormat("ERROR", str(message["msg"])))


class findFilePath:
    def __init__(self, filePath: dict) -> None:
        self.filePath = filePath["path"]
        self.sampleType = filePath["SampleType"]

    def getFilePath(self) -> dict:
        """从文件夹中筛选出需要分析的文件。

        :return: 分析所需的文件信息
        """
        # 输入校验
        if not self.valiDateSampleType():
            print("无效的样本类型。")
            return {"reg": 0, "msg": "无效的样本类型"}

        # 定义文件名常量
        PEAKTABLEFILE = "Peak Table.csv"
        SMEARTABLEFILE = "Smear Analysis Result.csv"
        QUALITYTABLEFILE = "Quality Table.csv"
        ELECTROPHEROGRAM = "Electropherogram.csv"
        pattern = r'\d{4}\s{1}\d{2}\s{1}\d{2}\s{1}\d{2}H\s{1}\d{2}M'

        peakTable = ""
        try:
            # 搜索文件并进行异常处理
            smearTable = self.findFileInPath(self.filePath, SMEARTABLEFILE)[0]
            if not smearTable:
                print(f"未找到 {SMEARTABLEFILE}。")
                return {"reg": 0, "msg": f"未找到 {SMEARTABLEFILE}"}

            qualityTable = self.findFileInPath(self.filePath, QUALITYTABLEFILE)[0]
            if not qualityTable:
                print(f"未找到 {QUALITYTABLEFILE}。")
                return {"reg": 0, "msg": f"未找到 {QUALITYTABLEFILE}"}

            if self.sampleType == "核酸":
                peakTable = self.findFileInPath(self.filePath, PEAKTABLEFILE)[0]
                if not peakTable:
                    print(f"未找到 {PEAKTABLEFILE}。")
                    return {"reg": 0, "msg": f"未找到 {PEAKTABLEFILE}"}
            match = re.search(pattern, smearTable)
            if match:
                if self.sampleType == "文库":
                    return {"reg": 1, "msg": {"smearTable": smearTable, "qualityTable": qualityTable, "saveName": match.group()}}
                elif self.sampleType == "核酸":
                    return {"reg": 1, "msg": {"smearTable": smearTable, "qualityTable": qualityTable, "peakTable": peakTable, "saveName": match.group()}}
        except Exception as e:
            # 对潜在的异常进行处理
            print(f"处理文件时发生错误：{e}")
            return {"reg": 0, "msg": e}

    def valiDateSampleType(self):
        """验证样本类型是否有效。

        :param sampletype: 样本类型
        :return: True 如果样本类型有效，否则 False
        """
        validTypes = {"核酸", "文库"}  # 定义有效样本类型集合
        return self.sampleType in validTypes

    def findFileInPath(self, path: str, filename: str) -> list[str]:
        """搜索指定路径下的文件。

        :param path: 文件夹路径
        :param filename: 文件名
        :return: 文件的完整路径列表
        """
        folderPath = path
        partialFileName = filename
        filePath = []
        a = f'{folderPath}/*{partialFileName}'
        for file in glob.iglob(f'{folderPath}/*{partialFileName}'):
            filePath.append(file)
        return filePath