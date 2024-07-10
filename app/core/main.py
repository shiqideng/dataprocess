# 标准库
import sys
import os
import re
import glob

# 第三方库
from PySide6.QtWidgets import QWidget, QFileDialog, QHeaderView
from PySide6.QtCore import QStandardPaths, Slot, Qt
from openpyxl import Workbook, load_workbook

# 本地包
from app.ui.Ui_MainWindow import Ui_Form
from app.lib import logger, caculat


class MainWindow(QWidget, Ui_Form):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.messagebox = None
        self.saveName = None
        self.analysis5400Status = False
        self.setupUi(self)
        self.run()

    def run(self):
        # 初始化配置
        self.getSystem()
        self.DataProcess5400ProgressBar.setValue(0)
        # 隐藏Quality标签页
        self.QTabWidget5400.removeTab(1)
        # 隐藏peaktable标签页
        self.QTabWidget5400.removeTab(1)
        # 隐藏labchip标签页
        self.QTabWidget5400.removeTab(1)

        self.ReName5400CheckBox.setChecked(False)

        # 使用lambda函数来传递参数
        self.Import5400FilePathToolButton.clicked.connect(lambda: self.selectFile("Import5400FilePathLineEdit"))
        self.Export5400FilePathToolButton.clicked.connect(lambda: self.selectFile("Export5400FilePathLineEdit"))
        self.Preview5400PushButton.clicked.connect(self.preview5400)
        self.Start5400PushButton.clicked.connect(self.start5400)
        self.Export5400PushButton.clicked.connect(self.exportToExcel)
        self.Clear5400PushButton.clicked.connect(self.clear5400)
        self.SampleType5400ComboBox.currentTextChanged.connect(self.handleIndexChanged)

    @Slot(str)
    def selectFile(self, lineEditName):
        defaultPath = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
        filePath = QFileDialog.getExistingDirectory(self, "选取文件夹", defaultPath)
        if filePath:
            getattr(self, lineEditName).setText(filePath)
        else:
            logger.logger.info("用户取消了文件夹选择操作或出现了错误。")
            self.logTextBrowser.append(caculat.logFormat("INFO", "用户取消了文件夹选择操作或出现了错误。"))

    def exportToExcel(self):
        if self.analysis5400Status:
            CELL_TO_CLEAR = ["备注", "空列", "原始质量浓度", "原始摩尔浓度"]
            MERGE_RANGE = "G1:H1"
            HEAD_TITLE = "片段大小"

            def get_file_path():
                dir_path = self.Export5400FilePathLineEdit.text()
                ensure_dir(dir_path)
                return {"reg": 1, "msg": {"exportResultPath": os.path.join(dir_path, f"{self.saveName}.xlsx"),
                                          "caculateResultPath": os.path.join(dir_path, f"{self.saveName}_backup.xlsx")}}

            def ensure_dir(path):
                """Ensure directory exists."""
                if not os.path.exists(path):
                    os.makedirs(path)

            # 获取文件路径
            filePath = get_file_path()
            if filePath["reg"] == 1:
                caculateResultPath = filePath["msg"]["caculateResultPath"]
                exportResultPath = filePath["msg"]["exportResultPath"]
                # 获取模型数据
                model = self.ResultTableAgilent5400TableView.model()
                columnCount = model.columnCount()
                rowCount = model.rowCount()

                # 创建工作簿
                wb = Workbook()
                sheet = wb.active

                # 写入表头
                for col in range(columnCount):
                    header = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
                    sheet.cell(row=1, column=col + 1).value = header

                # 写入内容
                for row in range(rowCount):
                    for col in range(columnCount):
                        value = model.data(model.index(row, col), Qt.DisplayRole)
                        sheet.cell(row=row + 2, column=col + 1).value = value
                wb.save(caculateResultPath)

                wb = load_workbook(caculateResultPath)
                sheet = wb.active
                # 删除孔号列
                sheet.delete_cols(1)

                # 清空指定单元格内容
                for row in sheet.iter_rows(min_row=0, max_row=1):
                    for cell in row:
                        if cell.value in CELL_TO_CLEAR:
                            cell.value = None

                # 合并单元格并设置标题
                sheet.merge_cells(range_string=MERGE_RANGE)
                sheet[MERGE_RANGE.split(":")[0]] = HEAD_TITLE

                # 清理特定列的0值
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    # 遍历第7至9列（索引为6到12）
                    for col_idx in range(6, 12):
                        # 获取单元格对象
                        cell = row[col_idx]
                        # 检查单元格的值是否为0
                        if cell.value == 0 or cell.value == "0":
                            # 清空单元格内容
                            cell.value = None

                # 将第3至5列单元格中数字从str转换为float
                # for row in sheet.iter_rows(min_row=2, min_col=2, max_col=4):
                #     for cell in row:
                #         if cell.value is not None and isinstance(cell.value, str):
                #             # 将字符串转换为浮点数(带小数的字符串无法直接转换为整数)
                #             value = float(cell.value)
                #             # 从浮点数转换为整数（四舍五入）
                #             cell.value = round(value)

                # 保存Excel文件
                wb.save(exportResultPath)

                # 更新日志和消息框
                log_message = caculat.logFormat("INFO",
                                               f'可上传数据已成功导出到{exportResultPath}!\n 分析数据已成功导出至{caculateResultPath}!\n 请检查数据是否正确。')
                self.logTextBrowser.append(log_message)
                logger.logger.info(log_message)
                self.messagebox = caculat.MessageBox(Icon="Information", text=log_message)
                self.messagebox.show()
            elif filePath["reg"] == 0:
                logger.logger.error(filePath["msg"])
                self.logTextBrowser.append(caculat.logFormat("ERROR", filePath["msg"]))
        else:
            self.messagebox = caculat.MessageBox(Icon="Warning", text="未分析数据，无法导出结果！")
            self.messagebox.show()

    def handleIndexChanged(self):
        if self.SampleType5400ComboBox.currentText() == "核酸":
            self.QTabWidget5400.insertTab(2, self.PeakTable5400tab, "Peak Table")
        elif self.SampleType5400ComboBox.currentText() == "文库":
            self.QTabWidget5400.removeTab(2)

    def start5400(self):
        if self.Import5400FilePathLineEdit.text():
            try:
                self.upDatePrograssBar(0)
                self.preview5400()
                self.QTabWidget5400.setCurrentIndex(1)
                self.upDatePrograssBar(15)
                filePath = findFilePath({"path": self.Import5400FilePathLineEdit.text(),
                                         "SampleType": self.SampleType5400ComboBox.currentText()})
                filePath = filePath.getFilePath()
                # 判断是否选中重命名对话框
                if self.ReName5400CheckBox.isChecked():
                    params = {
                        "input_folder": self.Import5400FilePathLineEdit.text(),
                        "output_folder": self.Export5400FilePathLineEdit.text(),
                    }
                    rename = caculat.reNameImage(params)
                    if rename["reg"] == 1:
                        self.logTextBrowser.append(caculat.logFormat("INFO", "图片重命名成功！"))
                        logger.logger.info("图片重命名成功！")
                        self.upDatePrograssBar(20)
                    elif rename["reg"] == 0:
                        self.logTextBrowser.append(caculat.logFormat("ERROR", rename["msg"]))
                        logger.logger.error(rename["msg"])

                if filePath["reg"] == 1:
                    self.saveName = filePath["msg"]["saveName"]
                    self.upDatePrograssBar(30)
                    self.logTextBrowser.append(caculat.logFormat("INFO", f"保存名称：{self.saveName}.xlsx"))
                    logger.logger.info(f"文库计算结果保存名称：{self.saveName}.xlsx")
                    if self.SampleType5400ComboBox.currentText() == "核酸":
                        pass
                    elif self.SampleType5400ComboBox.currentText() == "文库":
                        self.logTextBrowser.append(caculat.logFormat("INFO", "正在处理准备文库数据..."))
                        logger.logger.info("正在处理准备文库数据...")
                        global qualityTablePth
                        qualityTablePth = filePath["msg"]["qualityTable"]
                        global smearTablePath
                        smearTablePath = filePath["msg"]["smearTable"]
                        filePath = {"SampleType": "文库", "qualityTablepath": qualityTablePth,
                                    "smearTablepath": smearTablePath}
                        self.worker = caculat.caculateResult(filePath)
                        if self.worker["reg"] == 1:
                            self.upDatePrograssBar(40)
                            # Resultcaculat = self.worker.prograssChange.connect(self.upDatePrograssBar)
                            self.logTextBrowser.append(caculat.logFormat("INFO", "开始计算文库数据..."))
                            Result = caculat.createModel({"type": 1, "path": self.worker["msg"]})
                            logger.logger.info("文库数据计算完成")
                            self.logTextBrowser.append(caculat.logFormat("INFO", "文库数据计算完成"))
                            self.upDatePrograssBar(90)
                            # self.ResultTableLabChip5400TableView.setModel(LabChipResult)
                            self.ResultTableAgilent5400TableView.setModel(Result)
                            # 结果判定下拉列表选择
                            # judgeList = caculat.getConfig('conf/config.ini','Judge', 'judge')["msg"].split(",")
                            # self.ResultTableAgilent5400TableView.setItemDelegateForColumn(6,caculat.MyComboBoxDelegate(judgeList, self.ResultTableAgilent5400TableView))
                            self.upDatePrograssBar(99)
                            self.logTextBrowser.append(caculat.logFormat("INFO", "显示数据..."))
                            self.upDatePrograssBar(100)
                            self.messageBox = caculat.MessageBox(Icon="Information", text="计算完成")
                            self.analysis5400Status = True
                        elif self.worker["reg"] == 0:
                            self.messageBox = caculat.MessageBox(Icon="Critical", text="计算失败")
                            self.logTextBrowser.append(caculat.logFormat("ERROR", self.worker["msg"]))
                            logger.logger.error(self.worker["msg"])
                        # 设置表格行和列自适应
                        self.ResultTableLabChip5400TableView.horizontalHeader().setSectionResizeMode(
                            QHeaderView.Stretch)
                        self.ResultTableLabChip5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableLabChip5400TableView.resizeColumnsToContents()

                        self.ResultTableAgilent5400TableView.horizontalHeader().setSectionResizeMode(
                            QHeaderView.Stretch)
                        self.ResultTableAgilent5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.ResultTableAgilent5400TableView.resizeColumnsToContents()

                else:
                    self.logTextBrowser.append(caculat.logFormat("ERROR", str(filePath["msg"])))
                    logger.logger.error(str(filePath["msg"]))
            except IOError:
                self.logTextBrowser.append(caculat.logFormat("ERROR", "部分文件存在异常" + str(IOError)))
                logger.logger.error("部分文件存在异常" + str(IOError))
            except Exception as e:
                self.logTextBrowser.append(caculat.logFormat("ERROR", "发生意外错误" + str(e)))
                logger.logger.error("发生意外错误" + str(e))
        else:
            self.logTextBrowser.append(caculat.logFormat("ERROR", "请选择导入文件路径。"))
            logger.logger.error("未选择导入文件路径。")
        self.ReName5400CheckBox.setChecked(False)

    def clear5400(self):
        self.messagebox = caculat.MessageBox(Icon="Question", text="确定清空所有数据吗？")
        self.messagebox.resultReady.connect(self.handleResultReady)
        self.messagebox.show()

    def clearData(self):
        # 定义需要清除数据的表格视图列表
        tablesToClear = [
            self.PeakTable5400TableView,
            self.SmearTable5400TableView,
            self.QualityTable5400TableView,
            self.ResultTableAgilent5400TableView
        ]

        # 清除文本框中的数据
        self.Export5400FilePathLineEdit.clear()
        self.Import5400FilePathLineEdit.clear()

        # 遍历表格视图列表，尝试清除每个表格的数据
        for tableView in tablesToClear:
            # 为了提高代码健壮性，添加了对table_view对象存在性的检查
            if tableView is not None:
                # 再次检查模型是否存在，如果存在，则清除模型
                model = tableView.model()
                if model is not None:
                    tableView.setModel(None)
            else:
                # 可以选择在这里记录日志或者进行一些错误处理
                self.logTextBrowser.append(format("ERROR"), f"Warning: Table view {tableView} is None.")
                logger.logger.warning(f"Warning: Table view {tableView} is None.")

    def preview5400(self):
        if self.Import5400FilePathLineEdit.text():
            try:
                filePath = findFilePath({"path": self.Import5400FilePathLineEdit.text(),
                                         "SampleType": self.SampleType5400ComboBox.currentText()})
                filePath = filePath.getFilePath()
                if filePath["reg"] == 1:
                    self.saveName = filePath["msg"]["saveName"]
                    if self.SampleType5400ComboBox.currentText() == "核酸":
                        self.logTextBrowser.append(caculat.logFormat("INFO", "核酸分析暂未支持，请等待后续升级！"))
                        self.messageBox = caculat.MessageBox(Icon="Warning", text="核酸结果分析暂未支持，请等待后续升级！")
                        self.messageBox.show()

                        # self.logTextBrowser.append(caculat.logFormat("INFO", "qualityTablePath:" + str(filePath["msg"]["qualityTable"])))
                        # self.logTextBrowser.append(caculat.logFormat("INFO", "smearTablePath:" + str(filePath["msg"]["smearTable"])))
                        # self.logTextBrowser.append(caculat.logFormat("INFO", "peakTablePath:" + str(filePath["msg"]["peakTable"])))

                        # global qualityTablepath
                        # qualityTablePath = filePath["msg"]["qualityTable"]
                        # quality = caculat.createModel(qualityTablePath)
                        # self.QualityTable5400TableView.setModel(quality)
                        # # 设置表格行和列自适应
                        # self.QualityTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.QualityTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                        # global smearTablepath
                        # smearTablePath = filePath["msg"]["smearTable"]
                        # smear = caculat.createModel(smearTablePath)
                        # self.SmearTable5400TableView.setModel(smear)
                        # # 设置表格行和列自适应
                        # self.SmearTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.SmearTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                        # global peakTablePath
                        # peakTablePath = filePath["msg"]["peakTable"]
                        # peak =caculat.createModel(peakTablePath)
                        # self.PeakTable5400TableView.setModel(peak)
                        # # 设置表格行和列自适应
                        # self.PeakTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.PeakTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

                    elif self.SampleType5400ComboBox.currentText() == "文库":
                        # self.logTextBrowser.append(caculat.logFormat("INFO", "qualityTablePath:" + str(filePath["msg"]["qualityTable"])))
                        self.logTextBrowser.append(
                            caculat.logFormat("INFO", "smearTablePath:" + str(filePath["msg"]["smearTable"])))
                        logger.logger.info("smearTablePath:" + str(filePath["msg"]["smearTable"]))

                        # # 预览QualityTable
                        # global qualityTablepath
                        # qualityTablepath = filePath["msg"]["qualityTable"]
                        # quality = caculat.createModel({"type": 2, "path":qualityTablepath})
                        # self.QualityTable5400TableView.setModel(quality)
                        # # 设置表格行和列自适应
                        # self.QualityTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.QualityTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        # self.QualityTable5400TableView.resizeColumnsToContents()

                        global smearTablepath
                        smearTablepath = filePath["msg"]["smearTable"]
                        smear = caculat.createModel({"type": 2, "path": smearTablepath})
                        self.SmearTable5400TableView.setModel(smear)
                        # 设置表格行和列自适应
                        self.SmearTable5400TableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                        self.SmearTable5400TableView.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)


                else:
                    self.logTextBrowser.append(caculat.logFormat("ERROR", str(filePath["msg"])))
                    logger.logger.error(str(filePath["msg"]))
            except IOError:
                self.logTextBrowser.append(caculat.logFormat("ERROR", "部分文件存在异常" + str(IOError)))
                logger.logger.error("部分文件存在异常" + str(IOError))
            except Exception as e:
                self.logTextBrowser.append(caculat.logFormat("ERROR", "发生意外错误" + str(e)))
                logger.logger.error("发生意外错误" + str(e))
        else:
            self.logTextBrowser.append(caculat.logFormat("ERROR", "请选择导入文件路径。"))
            logger.logger.error("未选择导入文件路径。")

    def getSystem(self):
        try:
            if sys.platform == 'linux' or sys.platform == 'darwin':
                importPath = caculat.getConfig('conf/config.ini', 'Import Paths', 'linux_path')
                exportPath = caculat.getConfig('conf/config.ini', 'Export Paths', 'linux_path')
            elif sys.platform == 'win32':
                importPath = caculat.getConfig('conf/config.ini', 'Import Paths', 'win32_path')
                exportPath = caculat.getConfig('conf/config.ini', 'Export Paths', 'win32_path')
            # 使用 os.path.join 来构造路径，即使这里的路径已经是完整的
            importPath = os.path.join(importPath["msg"])
            exportPath = os.path.join(exportPath["msg"])
            # 判断是否读取到了配置文件中的路径
            if importPath != "No section: 'Import Paths'" and exportPath != "No section: 'Import Paths'":
                self.Import5400FilePathLineEdit.setText(importPath)
                self.Export5400FilePathLineEdit.setText(exportPath)
        except Exception as e:
            # 在这里处理可能的异常，比如读取配置文件失败、路径不存在等
            self.logTextBrowser.append(caculat.logFormat("ERROR", str(e)))
            logger.logger.error(str(e))

    def upDatePrograssBar(self, value):
        self.DataProcess5400ProgressBar.setValue(value)

    @Slot(dict)
    def handleResultReady(self, message):
        if message["reg"] == 1:
            self.logTextBrowser.append(caculat.logFormat("INFO", "用户确认清除内容"))
            logger.logger.info("用户确认清空内容")
            self.clearData()
            # self.logOperationSuccess()
        else:
            # self.logOperationCancelled()
            self.logTextBrowser.append(caculat.logFormat("ERROR", str(message["msg"])))
            logger.logger.error(str(message["msg"]))


class findFilePath:
    def __init__(self, filePath: dict) -> None:
        self.filePath = filePath["path"]
        self.sampleType = filePath["SampleType"]

    def getFilePath(self) -> dict:
        """从文件夹中筛选出需要分析的文件。

        :return: 分析所需的文件信息
        """
        # 输入校验
        if not self.valiDateSampleType():
            print("无效的样本类型。")
            return {"reg": 0, "msg": "无效的样本类型"}

        # 定义文件名常量
        PEAKTABLEFILE = "Peak Table.csv"
        SMEARTABLEFILE = "Smear Analysis Result.csv"
        QUALITYTABLEFILE = "Quality Table.csv"
        ELECTROPHEROGRAM = "Electropherogram.csv"
        pattern = r'\d{4}\s{1}\d{2}\s{1}\d{2}\s{1}\d{2}H\s{1}\d{2}M'

        peakTable = ""
        try:
            # 搜索文件并进行异常处理
            smearTable = self.findFileInPath(self.filePath, SMEARTABLEFILE)[0]
            if not smearTable:
                print(f"未找到 {SMEARTABLEFILE}。")
                return {"reg": 0, "msg": f"未找到 {SMEARTABLEFILE}"}

            # qualityTable = self.findFileInPath(self.filePath, QUALITYTABLEFILE)[0]
            # if not qualityTable:
            #     print(f"未找到 {QUALITYTABLEFILE}。")
            #     return {"reg": 0, "msg": f"未找到 {QUALITYTABLEFILE}"}

            # if self.sampleType == "核酸":
            #     peakTable = self.findFileInPath(self.filePath, PEAKTABLEFILE)[0]
            #     if not peakTable:
            #         print(f"未找到 {PEAKTABLEFILE}。")
            #         return {"reg": 0, "msg": f"未找到 {PEAKTABLEFILE}"}
            match = re.search(pattern, smearTable)
            if match:
                if self.sampleType == "文库":
                    return {"reg": 1, "msg": {"smearTable": smearTable, "qualityTable": "", "saveName": match.group()}}
                elif self.sampleType == "核酸":
                    return {"reg": 1, "msg": {"smearTable": smearTable, "qualityTable": "", "peakTable": peakTable,
                                              "saveName": match.group()}}
        except Exception as e:
            # 对潜在的异常进行处理
            return {"reg": 0, "msg": e}

    def valiDateSampleType(self):
        """验证样本类型是否有效。

        :param sampletype: 样本类型
        :return: True 如果样本类型有效，否则 False
        """
        validTypes = {"核酸", "文库"}  # 定义有效样本类型集合
        return self.sampleType in validTypes

    def findFileInPath(self, path: str, filename: str) -> list[str]:
        """搜索指定路径下的文件。

        :param path: 文件夹路径
        :param filename: 文件名
        :return: 文件的完整路径列表
        """
        folderPath = path
        partialFileName = filename
        filePath = []
        a = f'{folderPath}/*{partialFileName}'
        for file in glob.iglob(f'{folderPath}/*{partialFileName}'):
            filePath.append(file)
        return filePath
